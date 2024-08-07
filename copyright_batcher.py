import os
import json
import shutil
import logging
import h5py
from tqdm import tqdm
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError
from datetime import datetime, timezone
from dateutil.parser import parse as date_parse
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
from logging.handlers import RotatingFileHandler
import warnings
import threading

load_dotenv()

# Configure logging with rotating file handlers
log_file = "logs/batcher.log"
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
handler = RotatingFileHandler(log_file, maxBytes=10*1024*1024, backupCount=5)
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logging.getLogger('').addHandler(handler)

# Constants
BATCH_SIZE = 750
SUB_BATCH_SIZE = 150
MAX_DIMENSION = 360
DEFAULT_DATE = os.getenv("DEFAULT_DATE", "08/2024")

# Load environment variables
IMAGE_DIRECTORY_FOLDER = os.getenv("IMAGE_DIRECTORY_FOLDER")
METADATA_HDF5_PATH = os.getenv("METADATA_HDF5_PATH", "meta/metadata.hdf5")
OUTPUT_DIR = os.getenv("OUTPUT_DIRECTORY_FOLDER")
TEMPLATE_PATH = os.getenv("TEMPLATE_PATH")

# Check if environment variables are loaded correctly
if not IMAGE_DIRECTORY_FOLDER or not METADATA_HDF5_PATH or not OUTPUT_DIR or not TEMPLATE_PATH:
    raise ValueError("One or more environment variables are missing. Please check the .env file.")

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

if not os.path.exists(os.path.dirname(METADATA_HDF5_PATH)):
    os.makedirs(os.path.dirname(METADATA_HDF5_PATH), exist_ok=True)

# Create new metadata HDF5 file
def create_metadata_hdf5():
    logging.info(f"Creating a new metadata HDF5 file '{METADATA_HDF5_PATH}'.")
    metadata = {}
    json_files = [f for f in os.listdir(IMAGE_DIRECTORY_FOLDER) if f.endswith(".json")]
    with h5py.File(METADATA_HDF5_PATH, "w") as f:
        for json_file in json_files:
            json_path = os.path.join(IMAGE_DIRECTORY_FOLDER, json_file)
            try:
                with open(json_path, "r") as jf:
                    metadata_content = json.load(jf)
                    if isinstance(metadata_content, list):
                        metadata_content = metadata_content[0]  # Assuming the first item is the dictionary
                    if isinstance(metadata_content, dict):
                        f.create_dataset(json_file.replace(".json", ""), data=json.dumps(metadata_content))
                        metadata[json_file.replace(".json", "")] = metadata_content
                    else:
                        logging.warning(f"Unexpected metadata format in file: {json_path}")
            except Exception as e:
                logging.error(f"Error reading JSON file {json_path}: {e}")
    logging.info("Created metadata HDF5 file from JSON files")
    print(f"Created metadata for {len(metadata)} images")
    return metadata

# Load metadata from HDF5 file
def load_metadata_from_hdf5():
    metadata = {}
    with h5py.File(METADATA_HDF5_PATH, "r") as f:
        for image_name in f.keys():
            try:
                metadata_content = json.loads(f[image_name][()])
                if isinstance(metadata_content, list):
                    metadata_content = metadata_content[0]  # Assuming the first item is the dictionary
                if isinstance(metadata_content, dict):
                    metadata[image_name] = metadata_content
                else:
                    logging.warning(f"Unexpected metadata format in dataset: {image_name}")
            except Exception as e:
                logging.error(f"Error reading dataset {image_name} from HDF5 file: {e}")
    logging.info("Loaded metadata from HDF5 file")
    print(f"Loaded metadata for {len(metadata)} images")
    return metadata

def extract_date(metadata):
    dates = []
    date_tags = [
        "EXIF:DateTimeOriginal", "EXIF:CreateDate", "IPTC:DateCreated", 
        "IPTC:DigitalCreationDate", "XMP:CreateDate", "XMP:DateCreated"
    ]
    date_formats = [
        "%Y:%m:%d", "%Y:%m:%d %H:%M:%S", "%Y:%m:%d %H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S%z"
    ]

    for tag in date_tags:
        date_str = metadata.get(tag)
        if date_str:
            for date_format in date_formats:
                try:
                    date = datetime.strptime(date_str.split('.')[0], date_format)
                    # Make sure all dates are naive or convert to a common timezone
                    if date.tzinfo is None:
                        date = date.replace(tzinfo=timezone.utc)
                    dates.append(date)
                    logging.info(f"Extracted date from tag: {tag} -> {date.strftime('%m/%Y')}")
                    break
                except ValueError as e:
                    logging.debug(f"Failed to parse date {date_str} with format {date_format}: {e}")
                    continue

    if dates:
        latest_date = max(dates)
        logging.info(f"Using latest extracted date: {latest_date.strftime('%m/%Y')}")
        return latest_date.strftime("%m/%Y")
    else:
        logging.info(f"No valid date found in metadata, using default date {DEFAULT_DATE}")
        return DEFAULT_DATE

def resize_image(image_path, output_path):
    try:
        with Image.open(image_path) as img:
            logging.info(f"Opened image for resizing: {image_path}")
            img.thumbnail((MAX_DIMENSION, MAX_DIMENSION), Image.LANCZOS)
            img.save(output_path)
            logging.info(f"Resized image saved to {output_path}")
    except Exception as e:
        logging.error(f"Error resizing image {image_path}: {e}")
        raise

def process_image(image, batch_dir, worksheet, idx, metadata):
    img_path = os.path.join(IMAGE_DIRECTORY_FOLDER, image)
    logging.info(f"Processing image: {img_path}")
    image_metadata = metadata.get(image.replace('.jpg', ''), {})
    if not image_metadata:
        logging.warning(f"No metadata found for image: {image}")
    try:
        metadata_date = extract_date(image_metadata)
    except Exception as e:
        logging.error(f"Error extracting date for image {image}: {e}")
        return False

    resized_img_path = os.path.join(batch_dir, image)
    try:
        resize_image(img_path, resized_img_path)
    except FileNotFoundError as e:
        logging.error(f"File not found: {img_path}. Error: {e}")
        return False

    worksheet.cell(row=idx + 11, column=1).value = idx + 1
    worksheet.cell(row=idx + 11, column=2).value = image
    worksheet.cell(row=idx + 11, column=3).value = image
    worksheet.cell(row=idx + 11, column=4).value = metadata_date
    worksheet.cell(row=idx + 11, column=5).value = image
    worksheet.cell(row=idx + 11, column=6).value = ""
    return True

def process_batch(images, batch_number, metadata, year):
    batch_dir = os.path.join(OUTPUT_DIR, f"{year}_batch_{batch_number}")
    os.makedirs(batch_dir, exist_ok=True)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_image, image, batch_dir, ws, idx, metadata) for idx, image in enumerate(images)]
        for future in tqdm(as_completed(futures), total=len(futures), desc=f"Processing batch {batch_number} for {year}"):
            try:
                result = future.result()
                if not result:
                    logging.warning(f"Image processing failed for an image")
            except Exception as e:
                logging.error(f"Error processing image: {e}")

    output_xlsx = os.path.join(batch_dir, f"diversity_photos_batch_{batch_number}.xlsx")
    wb.save(output_xlsx)

    for j in range(0, len(images), SUB_BATCH_SIZE):
        sub_batch = images[j:j + SUB_BATCH_SIZE]
        sub_batch_dir = os.path.join(batch_dir, f"sub_batch_{j // SUB_BATCH_SIZE + 1}")
        os.makedirs(sub_batch_dir, exist_ok=True)
        for image in sub_batch:
            resized_img_path = os.path.join(batch_dir, image)
            if not os.path.exists(resized_img_path):
                logging.warning(f"Resized image not found: {resized_img_path}")
                continue
            try:
                shutil.copy(resized_img_path, os.path.join(sub_batch_dir, image))
                logging.info(f"Copied {resized_img_path} to {os.path.join(sub_batch_dir, image)}")
            except FileNotFoundError as e:
                logging.error(f"File not found during copy: {resized_img_path}. Error: {e}")

    logging.info(f"Processed batch {batch_number} for {year}")

def process_images():
    images = [f for f in os.listdir(IMAGE_DIRECTORY_FOLDER) if f.endswith(".jpg")]
    print(f"Found {len(images)} images")
    images.sort()

    metadata = create_metadata_hdf5()  # Always create new metadata HDF5 file

    # Group images by year
    images_by_year = {}
    for image in images:
        image_metadata = metadata.get(image.replace('.jpg', ''), {})
        date_str = extract_date(image_metadata)
        year = date_str.split("/")[1]
        if year not in images_by_year:
            images_by_year[year] = []
        images_by_year[year].append(image)

    # Process each group of images by year
    for year, year_images in images_by_year.items():
        for i in tqdm(range(0, len(year_images), BATCH_SIZE), desc=f"Processing batches for {year}"):
            batch = year_images[i:i + BATCH_SIZE]
            batch_number = i // BATCH_SIZE + 1
            process_batch(batch, batch_number, metadata, year)

if __name__ == "__main__":
    process_images()
